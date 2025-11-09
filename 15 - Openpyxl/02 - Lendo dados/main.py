from openpyxl import load_workbook

arquivo = load_workbook(r"15 - Openpyxl\02 - Lendo dados\planilha_funcionarios.xlsx")
planilha_funcionarios = arquivo["Funcionários"]
salario_anthony = planilha_funcionarios['D6']
print(salario_anthony.value)

salario_total = 0

linha_12 = planilha_funcionarios[12] # Tupla
for celula in linha_12:
    pass
    #print(celula.value)

linhas = planilha_funcionarios[2:7] # Tupla de tuplas
for tupla in linhas:
    #print("")
    for celula in tupla:
        #print(celula.value)
        pass

coluna_salario = planilha_funcionarios["D"]
for salario in coluna_salario:
    if type(salario.value) == int:
        salario_total+= salario.value
    if type(salario.value) == float:
        salario_total+= salario.value
    else:
        continue

print(f"O salario total é R${salario_total:.2f}")

for linha in planilha_funcionarios.iter_rows(values_only=True, min_row=2, max_row=20):
    nome, departamento, idade, salario, data_admissao = linha # desempacotamento
    print("-" * 20)
    print(f"""Nome: {nome}
Departamento: {departamento}
Idade: {idade}
Salário: {salario}
Data de adimissão: {data_admissao.strftime("%d/%m/%Y")}""")