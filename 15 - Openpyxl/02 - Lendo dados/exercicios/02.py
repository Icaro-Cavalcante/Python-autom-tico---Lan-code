# Exercício 2 – Leitura de colunas
# Percorra toda a coluna “Nota Final” e exiba somente os alunos com nota acima de 8.0.

from openpyxl import load_workbook

arquivo = load_workbook(r"15 - Openpyxl\02 - Lendo dados\exercicios\alunos.xlsx")
alunos = arquivo['Alunos']

for linha in alunos.iter_rows(values_only=True, min_row=2):
    nome, curso, idade, nota, data = linha
    if nota >= 8:
        print(nome)

