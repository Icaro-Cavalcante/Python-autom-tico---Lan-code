# Exercício 3 – Relatório geral
# Percorra todos os registros (exceto o cabeçalho) e imprima um relatório no seguinte formato:

# ALUNO: Maria Souza
# CURSO: Python
# IDADE: 29
# NOTA FINAL: 9.0
# MATRÍCULA: 15/02/2023
# -------------------------

from openpyxl import load_workbook

arquivo = load_workbook(r"15 - Openpyxl\02 - Lendo dados\exercicios\alunos.xlsx")
alunos = arquivo['Alunos']

for linha in alunos.iter_rows(values_only=True, min_row=2):
    aluno, curso, idade, nota, matricula = linha
    print(f"ALUNO: {aluno}\nCURSO: {curso}\nIDADE: {idade}\nNOTA FINAL: {nota}\nMATRÍCULA: {matricula}")
    print("-------------------------")