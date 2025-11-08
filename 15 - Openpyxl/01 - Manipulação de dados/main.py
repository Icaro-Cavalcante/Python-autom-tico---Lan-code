# pip install openpyxl - Para instalar a biblioteca

from openpyxl import Workbook
arquivo = Workbook()


planilha_atual = arquivo.active
planilha_atual.title = "Alunos"

# planilha_atual['A1'] = "Nome"
# planilha_atual['B1'] = "Matrícula"
# planilha_atual['C1'] = "Idade"

# planilha_atual['A2'] = "Icaro"
# planilha_atual['B2'] = 65820
# planilha_atual['C2'] = 19

# Mas também pode ser assim:

planilha_atual.append(["Nome", "Matrícula", "Idade"])
planilha_atual.append(["Icaro", 65820, 19])

# Criando uma nova planilha no mesmo arquivo:

arquivo.create_sheet("Professores")
planilha_prof = arquivo["Professores"]

planilha_prof.append(["Nome", "Matéria", "Salário"])
planilha_prof.append(["Carlos", "Inglês", 4961.73])

# Printar o nome das planilhas

print(arquivo.sheetnames)

# Salvar o Arquivo

arquivo.save(r"15 - Openpyxl\01 - Manipulação de dados\arquivo.xlsx")


