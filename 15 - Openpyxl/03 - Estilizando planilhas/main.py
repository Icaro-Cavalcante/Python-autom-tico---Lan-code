from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

arquivo = Workbook()
planilha = arquivo.active
planilha.title = "Alunos"

# Fonte
fonte_cabecalho = Font(bold=True, color="FFFFFF")

# Fundo
fundo_cabecalho = PatternFill(fgColor="424549", fill_type="solid")

# Bordas
fina = Side(style="thin")
borda = Border(right=fina, left=fina, top=fina, bottom=fina)

planilha.append(["Nome", "Matrícula", "Idade", "Saldo_RU", "Data_Matrícula"])
for celula in planilha[1]:
    celula.font = fonte_cabecalho
    celula.fill = fundo_cabecalho
    celula.border = borda

agora = datetime.now()

planilha.append(["Icaro", 1234, 19, 9, agora])
planilha.append(["João", 5678, 18, 12, agora])
planilha.append(["Pedro", 4235, 20, 21, agora])
for celula in planilha["D"]:
    if celula == 1:
        continue
    celula.alignment = Alignment(horizontal="left")
    celula.number_format = "R$ #,##0.00"

for celula in planilha["E"]:
    if celula == 1:
        continue
    celula.number_format = "DD/MM/YY"


arquivo.save(r"15 - Openpyxl\03 - Estilizando planilhas\Alunos.xlsx")